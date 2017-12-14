import arrow
import pickle
import os
import openpyxl


"""module to house the student class which 
can track the dates that students have arrived
and the time they have come in ."""


class student:
    time = arrow.now() # there is no need to  store time a million times

    def __init__(self, name, age, job):
        self.name = name  # student name
        self.age = age
        self.job = job
        self.daysnumber = 0
        self.date = student.time.date()
        self.time = student.time.time()
        self.checkintime = student.time
        self.checkouttime = student.time
        self.totaltime = self.checkouttime - self.checkintime

        self.checkedin = False
        self.checkedout = True

        self.folder = r"students/" + self.name.split(" ")[0].lower()  # fix the conflicting name issue
        self.savename = self.folder+"/save.bin"
        self.excelname = self.folder + "/excel.xlsx"  # donot forget to rename the excel file

        # these are student detail , which are static
        # written to the top of the excel file
        self.initialvalues = \
            {
                "A1": "name",
                "B1": self.name,
                "C1": "age",
                "D1": self.age,
                "A2": "job",
                "B2": self.job,

            }
        self.totals = \
            {
                "A3": "total time",
                "B3": self.totaltime,
                "C3": "number of days",
                "D3": self.daysnumber,

            }

        self.checkinvalues = \
            {
                "A4": self.date,
                "B4": self.checkintime.time(),
            }
        self.checkoutvalues = \
            {
                "C4": self.checkouttime.time(),
                "D4": self.checkouttime - self.checkintime
            }


        if not os.path.exists(self.folder):
            os.makedirs(self.folder)
        if not os.path.isfile(self.excelname):
            self.excelfile = openpyxl.Workbook()
            self.excelfile.save(self.excelname)
            self.excelfile.close()

        self.sheetname = "Sheet"

        self.writeinitialdata()

    # dont forget to find an efficent way to combine the move functions
    # found way do not forget to impiment
    #make everything into a list
    #that seems great
    def move_totals(self):
        self.totals = \
            {
                "A3": "total time",
                "B3": self.totaltime,
                "C3": "number of days",
                "D3": self.daysnumber,

            }

    def moveinwriter(self):
        x = []
        y = []
        for i in self.checkinvalues:
            x.append(i[0]+str(int(i[1])+1))
            y.append(i)

        for i in y:
            del self.checkinvalues[i]


        self.checkinvalues = \
            {
                x[0]: self.date,
                x[1]: self.checkintime.time(),
            }

    def moveoutwriter(self):
        x = []
        y = []
        for i in self.checkoutvalues:
            x.append(i[0] + str(int(i[1]) + 1))
            y.append(i)

        for i in y:
            del self.checkoutvalues[i]

        self.checkoutvalues[x[0]]= self.checkouttime.time()
        self.checkoutvalues[x[1]]= self.checkouttime-self.checkintime


    def writeinitialdata(self):
        """input the intial information into the excel file"""

        self.excelfile = openpyxl.load_workbook(self.excelname)
        sheet = self.excelfile.get_sheet_by_name(self.sheetname)
        for i in self.initialvalues:
            sheet[i].value = self.initialvalues[i]
        self.excelfile.save(self.excelname)
        self.excelfile.close()
        self.write_totals()

    def write_totals(self):
        self.move_totals()
        self.excelfile = openpyxl.load_workbook(self.excelname)
        sheet = self.excelfile.get_sheet_by_name(self.sheetname)
        for i in self.totals:
            sheet[i].value = self.totals[i]
        self.excelfile.save(self.excelname)
        self.excelfile.close()

    def checkin(self):
        """must be called each time checked in"""

        if self.checkedout:
            student.time = arrow.now()
            self.date = student.time.date()
            self.time = student.time.time()
            self.checkintime = student.time

            self.moveinwriter()

            self.excelfile = openpyxl.load_workbook(self.excelname)
            sheet = self.excelfile.get_sheet_by_name(self.sheetname)

            for i in self.checkinvalues:
                sheet[i].value = self.checkinvalues[i]
            self.excelfile.save(self.excelname)
            self.excelfile.close()

            self.checkedout = False
            self.checkedin = True

            savefile = open(self.savename,"wb")
            pickle.dump(self,savefile)
            savefile.close()
            self.write_totals()



    def checkout(self):
        """must be called each time checked out"""

        if self.checkedin:
            student.time = arrow.now()
            self.date = student.time.date()
            self.time = student.time.time()
            self.checkouttime = student.time
            self.daysnumber+=1
            self.totaltime = self.totaltime + (self.checkouttime - self.checkintime)
            self.moveoutwriter()

            self.excelfile = openpyxl.load_workbook(self.excelname)
            sheet = self.excelfile.get_sheet_by_name(self.sheetname)
            for i in self.checkoutvalues:
                sheet[i].value = self.checkoutvalues[i]
            self.excelfile.save(self.excelname)
            self.excelfile.close()

            self.checkedout = True
            self.checkedin = False

            savefile = open(self.savename, "wb")
            pickle.dump(self, savefile)
            savefile.close()
            self.write_totals()





